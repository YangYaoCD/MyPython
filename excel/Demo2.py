# coding=utf-8

"""
Author : YangYao
Date : 2020/5/31 10:25

xlsx文件的读写
"""

import openpyxl


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
    # print(sheet.rows)
    # print(list(sheet.rows)[0])
    # print(list(sheet.rows)[0][0].value)


def opera_xlsx(path1, path2, sheet_name1, sheet_name2, col):
    workbook2 = openpyxl.Workbook()
    sheet2 = workbook2.active
    sheet2.title = sheet_name2

    workbook1 = openpyxl.load_workbook(path1)
    sheet1 = workbook1[sheet_name1]

    print("\n复制：")
    print(sheet1.cell(row=1, column=1).value)
    print(sheet1.max_row)
    for i in range(sheet1.max_row):
        for j in range(sheet1.max_column):
            sheet2.cell(row=i + 1, column=j + 1).value = sheet1.cell(row=i + 1, column=j + 1).value

    sum=0.0
    for i in range(sheet1.max_row):
        if i!=0:
            sum=sum+float(sheet1.cell(row=i+1,column=3).value)
    sheet2.cell(row=sheet1.max_row + 1, column=3).value = sum

    workbook2.save(path2)
    print("xlsx格式表格写入数据成功！")


book_name_xlsx = 'xlsx格式测试工作簿.xlsx'
operator_name_xlsx = 'operator.xlsx'
sheet_name_xlsx = 'xlsx格式测试表'
value = [["姓名", "性别", "年龄", "城市", "职业"],
         ["111", "女", "66", "石家庄", "运维工程师"],
         ["222", "男", "55", "南京", "饭店老板"],
         ["333", "女", "27", "苏州", "保安"], ]

write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value)
read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)

opera_xlsx(book_name_xlsx, operator_name_xlsx, sheet_name_xlsx, sheet_name_xlsx, 1)
